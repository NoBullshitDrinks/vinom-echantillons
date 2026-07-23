#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Consolidation mensuelle des échantillons GDD2 -> 1 fichier Excel par domaine (tous agents cumulés),
envoyé à Yasmina PAR OUTLOOK via un flux Power Automate (aucun emailing tiers).

1. Récupère les lignes du mois depuis le Worker (GET /export).
2. Cumule les quantités par domaine/référence (toutes saisies d'agents confondues).
3. Génère l'Excel : Synthèse par domaine + Détail par domaine (facturation) + onglet analytique "Par agent".
4. POST le fichier (base64) vers un flux Power Automate qui l'envoie par Outlook.
   -> Appel serveur à serveur : pas de CORS, pas d'URL exposée dans le navigateur.

Variables d'environnement (secrets GitHub Actions) :
  WORKER_EXPORT_URL   ex. https://orange-math-f9f5.<sous-domaine>.workers.dev/export
  EXPORT_KEY          identique au secret EXPORT_KEY du Worker
  PA_EMAIL_WEBHOOK_URL  URL du flux Power Automate "Quand une requête HTTP est reçue"
  MAIL_TO             ex. adv@vinom.fr   (Yasmina)
  MAIL_CC             (optionnel) ex. a.durand@vinom.fr
Argument facultatif : le mois 'YYYY-MM' (défaut = mois précédent).
"""
import os, sys, base64, io, datetime, requests, pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def mois_precedent():
    t = datetime.date.today().replace(day=1) - datetime.timedelta(days=1)
    return t.strftime("%Y-%m")

MOIS = sys.argv[1] if len(sys.argv) > 1 and sys.argv[1] else mois_precedent()
EXPORT_URL = os.environ["WORKER_EXPORT_URL"]
EXPORT_KEY = os.environ["EXPORT_KEY"]

def fetch_rows():
    r = requests.get(EXPORT_URL, params={"mois": MOIS}, headers={"X-Export-Key": EXPORT_KEY}, timeout=30)
    r.raise_for_status()
    d = r.json()
    if not d.get("ok"): raise SystemExit("Export refusé : " + str(d))
    return d["rows"]

# ---------- styles ----------
HF=PatternFill('solid',start_color='1F4E79');HFo=Font(name='Arial',bold=True,size=10,color='FFFFFF');HA=Alignment('center','center',wrap_text=True)
SUB=PatternFill('solid',start_color='D9E1F2');SUBo=Font(name='Arial',bold=True,size=10)
TOT=PatternFill('solid',start_color='FFE699');TOTo=Font(name='Arial',bold=True,size=11)
MISS=PatternFill('solid',start_color='FCE4D6')   # ligne hors tarif : code article à identifier
CELL=Font(name='Arial',size=10);WR=Alignment(wrap_text=True,vertical='center')
TH=Side(border_style='thin',color='BFBFBF');BD=Border(TH,TH,TH,TH);EU='#,##0.00 €'
def sh(ws,r,n):
    for c in range(1,n+1):
        x=ws.cell(row=r,column=c);x.fill=HF;x.font=HFo;x.alignment=HA;x.border=BD
def sr(ws,r,n,fill=None,font=None,wrap=()):
    for c in range(1,n+1):
        x=ws.cell(row=r,column=c);x.font=font or CELL;x.border=BD
        if fill:x.fill=fill
        if c in wrap:x.alignment=WR

def build_excel(rows):
    df=pd.DataFrame(rows)
    if df.empty: raise SystemExit(f"Aucun échantillon pour {MOIS}.")
    df['qte']=pd.to_numeric(df['qte'],errors='coerce').fillna(0).astype(int)
    df['nf']=pd.to_numeric(df['nf'],errors='coerce')
    g=df.groupby(['nf','four','code','article','region'],dropna=False,as_index=False)['qte'].sum()
    SRC=f"Échantillons GDD2 – {MOIS} – Consolidé de tous les agents – quantités cumulées."
    wb=Workbook();wb.remove(wb.active)

    det=wb.create_sheet('2. Détail par fournisseur')
    det['A1']='DÉTAIL DES ÉCHANTILLONS À DEMANDER EN AVOIR – par domaine'
    det['A1'].font=Font(name='Arial',bold=True,size=14,color='1F4E79');det.merge_cells('A1:H1')
    det['A2']=SRC;det['A2'].font=Font(name='Arial',italic=True,size=9,color='595959');det.merge_cells('A2:H2')
    H=['N° Fourn.','Domaine / Fournisseur','Code article (Vinistoria)','Article','Région','Qté','Prix Brut HT','Montant avoir HT']
    for i,h in enumerate(H,1):det.cell(row=4,column=i,value=h)
    sh(det,4,len(H));r=5;drows=[]
    nfs=sorted(g['nf'].dropna().unique().tolist())+([float('nan')] if g['nf'].isna().any() else [])
    for nf in nfs:
        sub=g[g['nf'].isna()] if pd.isna(nf) else g[g['nf']==nf]
        sub=sub.sort_values('article');st=r
        for _,x in sub.iterrows():
            det.cell(row=r,column=1,value='' if pd.isna(x['nf']) else int(x['nf']))
            det.cell(row=r,column=2,value=x['four'])
            code=str(x['code']).strip()
            det.cell(row=r,column=3,value=code if code else "⚠ hors tarif – à identifier")
            det.cell(row=r,column=4,value=x['article']);det.cell(row=r,column=5,value=x['region'])
            det.cell(row=r,column=6,value=int(x['qte']))
            det.cell(row=r,column=8,value=f'=IF(G{r}="","",F{r}*G{r})')
            sr(det,r,len(H),fill=None if code else MISS,wrap=(4,))
            if code: det.cell(row=r,column=3).font=Font(name='Arial',size=10,bold=True)
            det.cell(row=r,column=7).number_format=EU;det.cell(row=r,column=8).number_format=EU
            drows.append(r);r+=1
        nom=sub.iloc[0]['four'] if len(sub) else ''
        det.cell(row=r,column=2,value=f"Sous-total – {nom}")
        det.cell(row=r,column=6,value=f'=SUM(F{st}:F{r-1})');det.cell(row=r,column=8,value=f'=SUM(H{st}:H{r-1})')
        sr(det,r,len(H),fill=SUB,font=SUBo);det.cell(row=r,column=8).number_format=EU;r+=1
    det.cell(row=r,column=2,value='TOTAL GÉNÉRAL')
    det.cell(row=r,column=6,value='=SUM('+','.join(f'F{i}'for i in drows)+')')
    det.cell(row=r,column=8,value='=SUM('+','.join(f'H{i}'for i in drows)+')')
    sr(det,r,len(H),fill=TOT,font=TOTo);det.cell(row=r,column=8).number_format=EU
    for i,w in enumerate([9,32,26,52,20,8,12,15],1):det.column_dimensions[get_column_letter(i)].width=w
    det.row_dimensions[4].height=26;det.freeze_panes='A5'

    syn=wb.create_sheet('1. Synthèse fournisseurs');wb.move_sheet('1. Synthèse fournisseurs',-(len(wb.sheetnames)-1))
    syn['A1']=f'AVOIRS À DEMANDER PAR DOMAINE – Échantillons GDD2 {MOIS}'
    syn['A1'].font=Font(name='Arial',bold=True,size=14,color='1F4E79');syn.merge_cells('A1:D1')
    syn['A2']=SRC;syn['A2'].font=Font(name='Arial',italic=True,size=9,color='595959');syn.merge_cells('A2:D2')
    H1=['N° Fourn.','Domaine / Fournisseur','Nb réfs','Nb bouteilles']
    for i,h in enumerate(H1,1):syn.cell(row=4,column=i,value=h)
    sh(syn,4,len(H1))
    s=g.groupby(['nf','four'],dropna=False,as_index=False).agg(refs=('code','nunique'),bts=('qte','sum')).sort_values('bts',ascending=False)
    r=5
    for _,x in s.iterrows():
        syn.cell(row=r,column=1,value='' if pd.isna(x['nf']) else int(x['nf']))
        syn.cell(row=r,column=2,value=x['four']);syn.cell(row=r,column=3,value=int(x['refs']));syn.cell(row=r,column=4,value=int(x['bts']))
        sr(syn,r,len(H1));r+=1
    syn.cell(row=r,column=2,value='TOTAL');syn.cell(row=r,column=3,value=f'=SUM(C5:C{r-1})');syn.cell(row=r,column=4,value=f'=SUM(D5:D{r-1})')
    sr(syn,r,len(H1),fill=TOT,font=TOTo)
    for i,w in enumerate([9,40,10,14],1):syn.column_dimensions[get_column_letter(i)].width=w
    syn.freeze_panes='A5'

    ag=wb.create_sheet('3. Par agent')
    ag['A1']=f'ANALYSE – Échantillons prélevés par agent ({MOIS})'
    ag['A1'].font=Font(name='Arial',bold=True,size=12,color='1F4E79');ag.merge_cells('A1:C1')
    for i,h in enumerate(['Agent','Nb références','Nb bouteilles'],1):ag.cell(row=3,column=i,value=h)
    sh(ag,3,3)
    a=df.groupby(['agent','agent_nom'],as_index=False).agg(refs=('code','count'),bts=('qte','sum')).sort_values('bts',ascending=False)
    r=4
    for _,x in a.iterrows():
        ag.cell(row=r,column=1,value=f"{x['agent_nom']} ({x['agent']})");ag.cell(row=r,column=2,value=int(x['refs']));ag.cell(row=r,column=3,value=int(x['bts']))
        sr(ag,r,3);r+=1
    for i,w in enumerate([34,14,14],1):ag.column_dimensions[get_column_letter(i)].width=w
    ag.freeze_panes='A4'

    bio=io.BytesIO();wb.save(bio);return bio.getvalue()

def send_via_powerautomate(xlsx_bytes):
    fname=f"Echantillons_GDD2_{MOIS}_consolide.xlsx"
    payload={
        "to": os.environ["MAIL_TO"],
        "cc": os.environ.get("MAIL_CC",""),
        "subject": f"Échantillons GDD2 à réclamer – {MOIS} (consolidé tous agents)",
        "body": (f"Bonjour Yasmina,<br><br>Ci-joint le relevé consolidé des échantillons prélevés à GDD2 "
                 f"en {MOIS}, tous agents confondus, regroupé par domaine.<br>"
                 f"Merci de réclamer les avoirs à chaque domaine.<br><br>Bien à toi."),
        "filename": fname,
        "contentBase64": base64.b64encode(xlsx_bytes).decode(),
    }
    r=requests.post(os.environ["PA_EMAIL_WEBHOOK_URL"], json=payload, timeout=60)
    r.raise_for_status()
    print("Power Automate:", r.status_code, "->", fname, "envoyé à", payload["to"])

if __name__=="__main__":
    rows=fetch_rows()
    print(f"{MOIS}: {len(rows)} lignes récupérées.")
    send_via_powerautomate(build_excel(rows))
    print("Fichier consolidé transmis à Power Automate (Outlook -> Yasmina).")
